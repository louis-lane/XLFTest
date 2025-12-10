import pandas as pd
from pathlib import Path
from typing import Tuple, Optional, List
from openpyxl.styles import PatternFill
from utils.core import CONFIG, log_errors, get_target_language, compress_ids, xliff_to_dataframe
from utils.glossary import get_glossary_map

def export_to_excel_with_glossary(root_path: Path, glossary_path: Optional[Path] = None) -> Tuple[int, int, int]:
    """
    Converts XLIFF files into optimized Excel files for translation.
    Deduplicates content and pre-translates using the glossary.

    Returns:
        Tuple[int, int, int]: (Files Processed, Languages Created, Error Count)
    """
    xliff_files = list(root_path.glob('*.xliff'))
    if not xliff_files:
        raise ValueError("No .xliff files found.")
    
    glossary_map = get_glossary_map(glossary_path)
            
    output_dir = root_path / str(CONFIG["folder_names"]["excel_export"])
    output_dir.mkdir(exist_ok=True)
    all_records: List[pd.DataFrame] = []
    errors: List[str] = []
    
    for file in xliff_files:
        try:
            lang = get_target_language(file)
            df = xliff_to_dataframe(file)
            if not df.empty:
                df['original_source_file'] = file.name
                df['language'] = lang
                all_records.append(df)
        except Exception as e:
            errors.append(f"Error reading {file.name}: {e}")
            
    if not all_records:
        if errors: log_errors(root_path, errors)
        raise ValueError("No content found.")
        
    master_df = pd.concat(all_records, ignore_index=True)
    master_df['source'] = master_df['source'].str.strip()
    processed_langs = 0
    
    for lang_code, lang_df in master_df.groupby('language'):
        try:
            master_path = output_dir / f"{lang_code}-master.xlsx"
            with pd.ExcelWriter(master_path, engine='openpyxl') as writer:
                sheet_name = f"{lang_code}-Translate_Here"
                
                # Deduplication Aggregation
                dedup = lang_df.groupby('source').agg(
                    existing_target=('existing_target', lambda x: next((s for s in x if s), '')),
                    count=('id', 'size'),
                    locations=('original_source_file', lambda x: ', '.join(x.unique())),
                    id_blob=('id', lambda x: compress_ids(list(x)))
                ).reset_index()

                dedup['target'] = dedup['existing_target']
                dedup['status'] = ''
                dedup['add_to_glossary'] = ''
                
                l_gloss = glossary_map.get(str(lang_code), {})
                
                for idx, row in dedup.iterrows():
                    src_lower = str(row['source']).lower()
                    if src_lower in CONFIG["protected_set"]:
                        dedup.at[idx, 'target'] = row['source']
                        dedup.at[idx, 'status'] = 'Protected'
                        continue
                    
                    if not row['target'] and row['source'] in l_gloss:
                        dedup.at[idx, 'target'] = l_gloss[row['source']]
                        dedup.at[idx, 'status'] = 'Glossary'
                    elif row['target']:
                        dedup.at[idx, 'status'] = 'Existing'
                
                dedup = dedup[['source', 'target', 'count', 'locations', 'status', 'add_to_glossary', 'id_blob']]
                dedup.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Formatting
                ws = writer.sheets[sheet_name]
                ws.column_dimensions['G'].hidden = True
                grey = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                for r_idx, status in enumerate(dedup['status'], start=2):
                    if status in ('Glossary', 'Protected'):
                        for c_idx in range(1, 8): ws.cell(row=r_idx, column=c_idx).fill = grey
                
                for fname, f_df in lang_df.groupby('original_source_file'):
                    s_name = Path(fname).stem[:31]
                    f_df.to_excel(writer, sheet_name=s_name, index=False)
                    writer.book[s_name].sheet_state = 'hidden'
            processed_langs += 1
        except Exception as e:
            errors.append(f"Error language {lang_code}: {e}")
            
    if errors: log_errors(root_path, errors)
    return len(xliff_files), processed_langs, len(errors)
