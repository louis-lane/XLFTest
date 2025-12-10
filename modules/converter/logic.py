import pandas as pd
from lxml import etree
from pathlib import Path
from openpyxl.styles import PatternFill
# UPDATED IMPORT:
from utils.core import CONFIG, log_errors, get_target_language, compress_ids, decompress_ids, xliff_to_dataframe
from utils.glossary import get_glossary_map, update_glossary_from_list

# ... (The rest of the file content remains exactly as it was in the previous step)
# Since the body didn't change, you can just update the import lines at the top.

def apply_deepl_translations(root_path, deepl_folder_path):
    master_folder = root_path / CONFIG["folder_names"]["excel_export"]
    if not master_folder.exists(): raise ValueError("Master folder not found.")

    if not deepl_folder_path: return 0, 0, ["User cancelled."]

    deepl_folder = Path(deepl_folder_path)
    master_files = list(master_folder.glob("*-master.xlsx"))
    deepl_files = list(deepl_folder.glob("*.xlsx"))

    if not master_files: raise ValueError("No master files found.")
    if not deepl_files: raise ValueError("No DeepL files found.")

    updated_count = 0
    errors = []

    for master_file in master_files:
        base_lang_code = master_file.name.replace("-master.xlsx", "")
        matching_deepl = next((df for df in deepl_files if df.name.lower().startswith(base_lang_code.lower())), None)
        
        if not matching_deepl:
            errors.append(f"No match for: {master_file.name}")
            continue

        try:
            deepl_df = pd.read_excel(matching_deepl, header=None, skiprows=1)
            translations = deepl_df.iloc[:, 0].astype(str).fillna('')
            master_wb = pd.ExcelFile(master_file)
            sheet_name = f"{base_lang_code}-Translate_Here"
            
            if sheet_name not in master_wb.sheet_names:
                errors.append(f"Sheet '{sheet_name}' missing in {master_file.name}")
                continue
                
            master_df = pd.read_excel(master_wb, sheet_name=sheet_name)
            if len(translations) != len(master_df):
                 errors.append(f"Row mismatch in {master_file.name}")
                 continue
                 
            master_df['target'] = translations
            with pd.ExcelWriter(master_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                master_df.to_excel(writer, sheet_name=sheet_name, index=False)
            updated_count += 1
        except Exception as e:
            errors.append(f"Error {master_file.name}: {e}")

    return updated_count, len(master_files), errors

def export_to_excel_with_glossary(root_path, glossary_path=None):
    xliff_files = list(root_path.glob('*.xliff'))
    if not xliff_files: raise ValueError("No .xliff files found.")
    
    glossary_map = get_glossary_map(glossary_path)
            
    output_dir = root_path / CONFIG["folder_names"]["excel_export"]
    output_dir.mkdir(exist_ok=True)
    all_records, errors = [], []
    
    for file in xliff_files:
        try:
            lang = get_target_language(file)
            df = xliff_to_dataframe(file)
            if not df.empty:
                df['original_source_file'] = file.name
                df['language'] = lang
                all_records.append(df)
        except Exception as e:
            errors.append(f"Error reading {file.name}: {e}")
            
    if not all_records:
        if errors: log_errors(root_path, errors)
        raise ValueError("No content found.")
        
    master_df = pd.concat(all_records, ignore_index=True)
    master_df['source'] = master_df['source'].str.strip()
    processed_langs = 0
    
    for lang_code, lang_df in master_df.groupby('language'):
        try:
            master_path = output_dir / f"{lang_code}-master.xlsx"
            with pd.ExcelWriter(master_path, engine='openpyxl') as writer:
                sheet_name = f"{lang_code}-Translate_Here"
                dedup = lang_df.groupby('source').agg(
                    existing_target=('existing_target', lambda x: next((s for s in x if s), '')),
                    count=('id', 'size'),
                    locations=('original_source_file', lambda x: ', '.join(x.unique())),
                    id_blob=('id', lambda x: compress_ids(list(x)))
                ).reset_index()

                dedup['target'] = dedup['existing_target']
                dedup['status'] = ''
                dedup['add_to_glossary'] = ''
                
                l_gloss = glossary_map.get(lang_code, {})
                
                for idx, row in dedup.iterrows():
                    src_lower = row['source'].lower()
                    if src_lower in CONFIG["protected_set"]:
                        dedup.at[idx, 'target'] = row['source']
                        dedup.at[idx, 'status'] = 'Protected'
                        continue
                    if not row['target'] and row['source'] in l_gloss:
                        dedup.at[idx, 'target'] = l_gloss[row['source']]
                        dedup.at[idx, 'status'] = 'Glossary'
                    elif row['target']:
                        dedup.at[idx, 'status'] = 'Existing'
                
                dedup = dedup[['source', 'target', 'count', 'locations', 'status', 'add_to_glossary', 'id_blob']]
                dedup.to_excel(writer, sheet_name=sheet_name, index=False)
                
                ws = writer.sheets[sheet_name]
                ws.column_dimensions['G'].hidden = True
                grey = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                for r_idx, status in enumerate(dedup['status'], start=2):
                    if status in ('Glossary', 'Protected'):
                        for c_idx in range(1, 8): ws.cell(row=r_idx, column=c_idx).fill = grey
                
                for fname, f_df in lang_df.groupby('original_source_file'):
                    s_name = Path(fname).stem[:31]
                    f_df.to_excel(writer, sheet_name=s_name, index=False)
                    writer.book[s_name].sheet_state = 'hidden'
            processed_langs += 1
        except Exception as e:
            errors.append(f"Error language {lang_code}: {e}")
            
    if errors: log_errors(root_path, errors)
    return len(xliff_files), processed_langs, len(errors)

def import_and_reconstruct_with_glossary(root_path, glossary_path=None):
    input_dir = root_path / CONFIG["folder_names"]["excel_export"]
    if not input_dir.exists(): raise ValueError("Excel export folder not found.")
    master_files = list(input_dir.glob('*-master.xlsx'))
    if not master_files: raise ValueError("No master files found.")
    
    errors = []
    try:
        new_entries = []
        for mf in master_files:
            lc = mf.name.replace("-master.xlsx", "")
            try:
                df = pd.read_excel(mf, sheet_name=f"{lc}-Translate_Here")
                if 'add_to_glossary' in df.columns:
                    for _, row in df[df['add_to_glossary'].astype(str).str.lower().isin(['x', 'yes'])].iterrows():
                        if pd.notna(row['target']):
                            new_entries.append({'source_text': row['source'], 'target_text': str(row['target']), 'language_code': lc})
            except: pass
        if new_entries and glossary_path: update_glossary_from_list(glossary_path, new_entries)
    except Exception as e: errors.append(f"Glossary update failed: {e}")

    trans_map = {}
    for mf in master_files:
        try:
            lc = mf.name.replace("-master.xlsx", "")
            if lc not in trans_map: trans_map[lc] = {}
            df = pd.read_excel(mf, sheet_name=f"{lc}-Translate_Here").dropna(subset=['target', 'id_blob'])
            for _, row in df.iterrows():
                tgt = str(row['target'])
                for uid in decompress_ids(str(row['id_blob'])):
                    trans_map[lc][uid] = tgt
        except Exception as e: errors.append(f"Map error {mf.name}: {e}")

    out_dir = root_path / CONFIG["folder_names"]["xliff_output"]
    sep_dir = out_dir / "Separate Languages"
    out_dir.mkdir(exist_ok=True); sep_dir.mkdir(exist_ok=True)
    
    cnt = 0
    for xf in list(root_path.glob('*.xliff')):
        try:
            lc = get_target_language(xf)
            l_map = trans_map.get(lc, {})
            if not l_map:
                for k in trans_map:
                    if k.lower() == lc.lower(): l_map = trans_map[k]; break
            
            tree = etree.parse(xf)
            ns = {'xliff': 'urn:oasis:names:tc:xliff:document:1.2'}
            for tu in tree.xpath('//xliff:trans-unit', namespaces=ns):
                uid = tu.get('id')
                if uid in l_map:
                    tn = tu.find('xliff:target', namespaces=ns)
                    if tn is None: tn = etree.SubElement(tu, f"{{{ns['xliff']}}}target")
                    tn.text = l_map[uid]
                    tn.set('state', 'translated')
            
            tree.write(out_dir / xf.name, pretty_print=True, xml_declaration=True, encoding='UTF-8')
            (sep_dir / lc).mkdir(exist_ok=True)
            tree.write(sep_dir / lc / xf.name, pretty_print=True, xml_declaration=True, encoding='UTF-8')
            cnt += 1
        except Exception as e: errors.append(f"Reconstruct error {xf.name}: {e}")

    repo = Path(CONFIG["folder_names"]["master_repo"])
    if errors: log_errors(root_path, errors)
    return cnt, len(errors)

def perform_analysis(root_path, glossary_path=None):
    xliff_files = list(root_path.glob('*.xliff'))
    if not xliff_files: raise ValueError("No XLIFF files.")
    
    g_map = get_glossary_map(glossary_path)

    records = []
    for f in xliff_files:
        df = xliff_to_dataframe(f)
        if not df.empty:
            df['language'] = get_target_language(f)
            df['wc'] = df['source'].str.split().str.len()
            records.append(df)
            
    if not records: raise ValueError("No content.")
    mdf = pd.concat(records)
    
    results = {}
    for lc, ldf in mdf.groupby('language'):
        uniq = ldf.drop_duplicates(subset=['source'])
        reps = ldf[ldf.duplicated(subset=['source'], keep=False)]
        rep_w = reps['wc'].sum() - uniq[uniq['source'].isin(reps['source'])]['wc'].sum()
        
        lg = g_map.get(lc, {})
        match_w = uniq[uniq['source'].isin(lg.keys())]['wc'].sum()
        new_w = uniq[~uniq['source'].isin(lg.keys())]['wc'].sum()
        
        results[lc] = {'Total Words': ldf['wc'].sum(), 'Repetitions': rep_w, 'Glossary Matches': match_w, 'New Words': new_w}
    return results
