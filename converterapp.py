import tkinter as tk
from tkinter import filedialog, messagebox
# Import standard ttk as 'tk_ttk' to access PanedWindow
from tkinter import ttk as tk_ttk 
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import pandas as pd
from lxml import etree
from pathlib import Path
import re
from openpyxl.styles import PatternFill
from datetime import datetime
import os
import zlib
import base64
import json
import sys
import threading

# --- DEFAULT CONFIGURATION (Fallback) ---
DEFAULT_CONFIG = {
    "folder_names": {
        "excel_export": "1_Excel_for_Translation",
        "xliff_output": "2_Translated_XLIFFs",
        "master_repo": "master_localization_files"
    },
    "protected_languages": [
        "English", "British English", "American English",
        "Español", "Spanish",
        "Français", "French",
        "Italiano", "Italian",
        "Deutsch", "German",
        "Português", "Portuguese", "Português (Brasil)",
        "Svenska", "Swedish",
        "Nederlands-Vlaamse", "Nederlands", "Dutch", "Vlaams",
        "Dansk", "Danish",
        "Norsk", "Norwegian",
        "Suomi", "Finnish",
        "Íslenska", "Icelandic",
        "Gaeilge", "Irish",
        "Cymraeg", "Welsh",
        "Polskie", "Polski", "Polish",
        "Čeština", "Czech",
        "Lietuvių", "Lithuanian",
        "Eesti", "Estonian",
        "Slovenčina", "Slovak",
        "Slovenščina", "Slovenian",
        "Magyar", "Hungarian",
        "Română", "Romanian",
        "Latviešu", "Latvian",
        "Hrvatski", "Croatian",
        "Srpski", "Serbian",
        "Bosanski", "Bosnian",
        "Русский", "Russian",
        "Українська", "Ukrainian",
        "Български", "Bulgarian",
        "Ελληνικά", "Greek",
        "Türk", "Türkçe", "Turkish",
        "Bahasa Indonesia", "Indonesian",
        "Bahasa Melayu", "Malay",
        "Tiếng Việt", "Vietnamese",
        "Tagalog", "Filipino",
        "ไทย", "Thai",
        "简体中文 (中国)", "简体中文", "Chinese (Simplified)",
        "繁體中文 (香港)", "繁體中文", "Chinese (Traditional)", "中國傳統 (香港)",
        "日本語", "Japanese",
        "한국어", "Korean",
        "العربية", "Arabic",
        "עברית", "Hebrew",
        "فارسی", "Persian",
        "اردو", "Urdu",
        "සිංහල", "Sinhala",
        "नेपाली", "Nepali",
        "বাংলা", "Bengali",
        "ગુજરાતી", "Gujarati",
        "हिंदी", "Hindi", "हिन्दी",
        "ಕನ್ನಡ", "Kannada",
        "മലയാളം", "Malayalam",
        "मराठी", "Marathi",
        "ଓଡ଼ିଆ", "Odia",
        "தமிழ்", "Tamil",
        "తెలుగు", "Telugu"
    ]
}

# --- CONFIG LOADER ---
def load_config():
    """Loads config.json if present; otherwise returns defaults."""
    if getattr(sys, 'frozen', False):
        application_path = Path(sys.executable).parent
    else:
        application_path = Path(__file__).parent

    config_path = application_path / "config.json"
    current_config = DEFAULT_CONFIG.copy()

    if config_path.exists():
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                user_config = json.load(f)
                if "protected_languages" in user_config:
                    current_config["protected_languages"] = user_config["protected_languages"]
                if "folder_names" in user_config:
                    current_config["folder_names"].update(user_config["folder_names"])
        except Exception as e:
            print(f"Warning: Could not load config.json. Using defaults. Error: {e}")

    current_config["protected_set"] = {x.lower() for x in current_config["protected_languages"]}
    return current_config

CONFIG = load_config()

# --- HELPER FUNCTIONS ---

def get_target_language(xliff_path):
    try:
        tree = etree.parse(xliff_path)
        file_node = tree.find('.//xliff:file', namespaces={'xliff': 'urn:oasis:names:tc:xliff:document:1.2'})
        return file_node.get('target-language', 'unknown') if file_node is not None else 'unknown'
    except Exception:
        return 'unknown'

def compress_ids(id_list):
    if not id_list: return ""
    try:
        full_string = "|".join(str(x) for x in id_list)
        compressed_data = zlib.compress(full_string.encode('utf-8'))
        return base64.b64encode(compressed_data).decode('utf-8')
    except Exception:
        return ""

def decompress_ids(blob_string):
    if not blob_string or pd.isna(blob_string) or str(blob_string).strip() == "":
        return []
    try:
        compressed_data = base64.b64decode(str(blob_string))
        full_string = zlib.decompress(compressed_data).decode('utf-8')
        return full_string.split('|')
    except Exception:
        return []

def log_errors_to_file(root_path, errors):
    log_path = Path(root_path) / "error_log.txt"
    with log_path.open("a", encoding="utf-8") as f:
        f.write(f"\n--- Log Entry: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ---\n")
        for error in errors:
            f.write(f"- {error}\n")

def update_glossary_file(glossary_path, new_entries):
    glossary_df = pd.read_excel(glossary_path) if Path(glossary_path).exists() else pd.DataFrame(columns=['source_text', 'target_text', 'language_code'])
    new_entries_df = pd.DataFrame(new_entries)
    combined_df = pd.concat([glossary_df, new_entries_df], ignore_index=True)
    combined_df.drop_duplicates(subset=['source_text', 'language_code'], keep='first', inplace=True)
    combined_df.to_excel(glossary_path, index=False)

def xliff_to_dataframe(xliff_path):
    ns = {'xliff': 'urn:oasis:names:tc:xliff:document:1.2'}
    tree = etree.parse(xliff_path)
    records = []
    non_translatable_strings = {'left-alignc', 'imagestatic-1', 'video2-1', 'imagehotspot-1'}
    non_translatable_patterns = [re.compile(r'text-\d+', re.IGNORECASE)]
    
    for trans_unit in tree.xpath('//xliff:trans-unit', namespaces=ns):
        if trans_unit.get('translate') == 'no': continue
        source_element = trans_unit.find('xliff:source', namespaces=ns)
        source_text = (source_element.text or '').strip()
        target_element = trans_unit.find('xliff:target', namespaces=ns)
        target_text = (target_element.text or '').strip() if target_element is not None else ''

        if not source_text: continue
        lower_source = source_text.lower()
        if lower_source.endswith(('.jpg', '.png')) or lower_source in non_translatable_strings or any(p.fullmatch(lower_source) for p in non_translatable_patterns):
            continue
        records.append({
            'id': trans_unit.get('id'),
            'source': source_text,
            'existing_target': target_text,
            'gomo-id (context)': trans_unit.get('gomo-id', '')
        })
    return pd.DataFrame(records)

def update_xliff_from_map(original_path, lang_specific_map):
    tree = etree.parse(original_path)
    ns = {'xliff': 'urn:oasis:names:tc:xliff:document:1.2'}
    etree.register_namespace('xliff', ns['xliff'])
    for trans_unit in tree.xpath('//xliff:trans-unit', namespaces=ns):
        unit_id = trans_unit.get('id')
        if unit_id and unit_id in lang_specific_map:
            target_element = trans_unit.find('xliff:target', namespaces=ns)
            if target_element is None:
                target_element = etree.SubElement(trans_unit, f"{{{ns['xliff']}}}target")
            target_element.text = lang_specific_map[unit_id]
            target_element.set('state', 'translated')
    return tree

# --- CORE LOGIC FUNCTIONS ---

def apply_deepl_translations(root_path):
    master_folder_name = CONFIG["folder_names"]["excel_export"]
    master_folder = root_path / master_folder_name
    
    if not master_folder.exists():
        raise ValueError(f"The '{master_folder_name}' folder was not found.")

    deepl_folder_path = filedialog.askdirectory(title="Select the Folder Containing DeepL Translated Files")
    if not deepl_folder_path:
        return 0, 0, ["User cancelled the operation."]

    deepl_folder = Path(deepl_folder_path)
    master_files = list(master_folder.glob("*-master.xlsx"))
    deepl_files = list(deepl_folder.glob("*.xlsx"))

    if not master_files:
        raise ValueError(f"No master Excel files found in '{master_folder_name}'.")
    if not deepl_files:
        raise ValueError(f"No .xlsx files found in the selected DeepL folder.")

    updated_count = 0
    errors = []

    for master_file in master_files:
        base_lang_code = master_file.name.replace("-master.xlsx", "")
        matching_deepl_file = None
        for df in deepl_files:
            if df.name.lower().startswith(base_lang_code.lower()):
                matching_deepl_file = df
                break
        
        if not matching_deepl_file:
            errors.append(f"No matching DeepL file found for master file: {master_file.name}")
            continue

        try:
            deepl_df = pd.read_excel(matching_deepl_file, header=None, skiprows=1)
            translations = deepl_df.iloc[:, 0].astype(str).fillna('')
            master_wb = pd.ExcelFile(master_file)
            sheet_name = f"{base_lang_code}-Translate_Here"
            if sheet_name not in master_wb.sheet_names:
                errors.append(f"Sheet '{sheet_name}' not found in {master_file.name}")
                continue
            master_df = pd.read_excel(master_wb, sheet_name=sheet_name)
            if len(translations) != len(master_df):
                 errors.append(f"Row count mismatch for {master_file.name}. Master has {len(master_df)} rows, DeepL file has {len(translations)} rows.")
                 continue
            master_df['target'] = translations
            with pd.ExcelWriter(master_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                master_df.to_excel(writer, sheet_name=sheet_name, index=False)
            updated_count += 1
        except Exception as e:
            errors.append(f"Failed to process {master_file.name} with {matching_deepl_file.name}: {e}")

    return updated_count, len(master_files), errors

def process_standard_files(root_path, translated_dir):
    repo_name = CONFIG["folder_names"]["master_repo"]
    repo_path = Path(repo_name)
    if not repo_path.exists():
        if getattr(sys, 'frozen', False):
            repo_path = Path(sys.executable).parent / repo_name
        if not repo_path.exists():
            return 

    prefixes = ["localization-localization_", "localizationerrors-localizationerrors_"]
    course_id = None
    all_translated_files = list(translated_dir.glob('*.xliff'))
    content_files = [f for f in all_translated_files if not f.name.startswith(('localization-', 'localizationerrors-'))]
    
    if content_files:
        try:
            tree = etree.parse(str(content_files[0]))
            file_node = tree.find('.//xliff:file', namespaces={'xliff': 'urn:oasis:names:tc:xliff:document:1.2'})
            if file_node is not None:
                course_id = file_node.get('id')
        except Exception as e:
            raise ValueError(f"Could not determine course ID from content files: {e}")
    if not course_id: return

    for prefix in prefixes:
        for standard_file in translated_dir.glob(f"{prefix}*.xliff"):
            lang_code = standard_file.name.replace(prefix, "").replace(".xliff", "")
            master_file_path = repo_path / standard_file.name
            if master_file_path.exists():
                try:
                    master_tree = etree.parse(str(master_file_path))
                    file_node = master_tree.find('.//xliff:file', namespaces={'xliff': 'urn:oasis:names:tc:xliff:document:1.2'})
                    if file_node is not None:
                        file_node.set('id', course_id)
                    master_tree.write(str(standard_file), pretty_print=True, xml_declaration=True, encoding='UTF-8')
                    sorted_lang_path = translated_dir / "Separate Languages" / lang_code / standard_file.name
                    if sorted_lang_path.exists():
                        master_tree.write(str(sorted_lang_path), pretty_print=True, xml_declaration=True, encoding='UTF-8')
                except Exception as e:
                    log_errors_to_file(root_path, [f"Failed to replace master file {standard_file.name}: {e}"])

def perform_analysis(root_path, glossary_path=None):
    xliff_files = list(root_path.glob('*.xliff'))
    if not xliff_files: raise ValueError("No .xliff files were found.")
    glossary_map = {}
    if glossary_path and Path(glossary_path).exists():
        glossary_df = pd.read_excel(glossary_path)
        for _, row in glossary_df.iterrows():
            lang = row.get('language_code', 'unknown')
            if lang not in glossary_map: glossary_map[lang] = {}
            glossary_map[lang][row['source_text'].strip()] = row['target_text']
    all_records = []
    for file in xliff_files:
        lang = get_target_language(file)
        df = xliff_to_dataframe(file)
        if not df.empty:
            df['language'] = lang
            df['word_count'] = df['source'].str.split().str.len()
            all_records.append(df)
    if not all_records: raise ValueError("No translatable content found to analyze.")
    master_df = pd.concat(all_records, ignore_index=True)
    master_df['source'] = master_df['source'].str.strip()
    analysis_results = {}
    for lang_code, lang_df in master_df.groupby('language'):
        unique_segments = lang_df.drop_duplicates(subset=['source'], keep='first')
        repeated_segments = lang_df[lang_df.duplicated(subset=['source'], keep=False)]
        reps_words = repeated_segments['word_count'].sum() - unique_segments[unique_segments['source'].isin(repeated_segments['source'])]['word_count'].sum()
        lang_glossary = glossary_map.get(lang_code, {})
        is_glossary_match = unique_segments['source'].isin(lang_glossary.keys())
        glossary_words = unique_segments[is_glossary_match]['word_count'].sum()
        is_new = ~unique_segments['source'].isin(lang_glossary.keys())
        new_words = unique_segments[is_new]['word_count'].sum()
        analysis_results[lang_code] = {'Total Words': lang_df['word_count'].sum(), 'Repetitions': reps_words, 'Glossary Matches': glossary_words, 'New Words': new_words}
    return analysis_results

def export_to_excel_with_glossary(root_path, glossary_path=None):
    xliff_files = list(root_path.glob('*.xliff'))
    if not xliff_files: raise ValueError("No .xliff files were found.")
    glossary_map = {}
    if glossary_path and Path(glossary_path).exists():
        glossary_df = pd.read_excel(glossary_path)
        for _, row in glossary_df.iterrows():
            lang = row.get('language_code', 'unknown')
            if lang not in glossary_map: glossary_map[lang] = {}
            glossary_map[lang][row['source_text'].strip()] = row['target_text']
            
    output_dir_name = CONFIG["folder_names"]["excel_export"]
    output_dir = root_path / output_dir_name
    output_dir.mkdir(exist_ok=True)
    all_records, errors = [], []
    
    for file in xliff_files:
        try:
            lang = get_target_language(file)
            df = xliff_to_dataframe(file)
            if not df.empty:
                df['original_source_file'] = file.name
                df['language'] = lang
                all_records.append(df)
        except Exception as e:
            errors.append(f"Failed to read or process {file.name}: {e}")
            
    if not all_records:
        if errors: log_errors_to_file(root_path, errors)
        raise ValueError("No translatable content was found after applying filters.")
        
    master_df = pd.concat(all_records, ignore_index=True)
    master_df['source'] = master_df['source'].str.strip()
    processed_langs = 0
    
    for lang_code, lang_df in master_df.groupby('language'):
        try:
            master_workbook_path = output_dir / f"{lang_code}-master.xlsx"
            with pd.ExcelWriter(master_workbook_path, engine='openpyxl') as writer:
                translate_sheet_name = f"{lang_code}-Translate_Here"
                deduplicated = lang_df.groupby('source').agg(
                    existing_target=('existing_target', lambda x: next((s for s in x if s), '')),
                    count=('id', 'size'),
                    locations=('original_source_file', lambda x: ', '.join(x.unique())),
                    id_blob=('id', lambda x: compress_ids(list(x)))
                ).reset_index()

                deduplicated['target'] = deduplicated['existing_target']
                deduplicated['status'] = ''
                deduplicated['add_to_glossary'] = ''
                lang_glossary = glossary_map.get(lang_code, {})
                for index, row in deduplicated.iterrows():
                    source_lower = row['source'].lower()
                    if source_lower in CONFIG["protected_set"]:
                        deduplicated.at[index, 'target'] = row['source'] 
                        deduplicated.at[index, 'status'] = 'Protected (Language Name)'
                        continue 
                    if not row['target'] and row['source'] in lang_glossary:
                        deduplicated.at[index, 'target'] = lang_glossary[row['source']]
                        deduplicated.at[index, 'status'] = 'Pre-translated from Glossary'
                    elif row['target']:
                        deduplicated.at[index, 'status'] = 'Existing translation'
                
                deduplicated = deduplicated[['source', 'target', 'count', 'locations', 'status', 'add_to_glossary', 'id_blob']]
                deduplicated.to_excel(writer, sheet_name=translate_sheet_name, index=False)
                workbook, worksheet = writer.book, writer.sheets[translate_sheet_name]
                worksheet.column_dimensions['G'].hidden = True
                grey_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                for row_idx, status in enumerate(deduplicated['status'], start=2):
                    if status in ('Pre-translated from Glossary', 'Protected (Language Name)'):
                        for col_idx in range(1, len(deduplicated.columns) + 1):
                            worksheet.cell(row=row_idx, column=col_idx).fill = grey_fill
                
                for filename, file_df in lang_df.groupby('original_source_file'):
                    sheet_name = Path(filename).stem[:31]
                    if 'existing_target' in file_df.columns:
                        file_df.to_excel(writer, sheet_name=sheet_name, index=False, columns=['id', 'source', 'existing_target', 'gomo-id (context)'])
                    else:
                         file_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    workbook[sheet_name].sheet_state = 'hidden'
                workbook.active = worksheet
            processed_langs += 1
        except Exception as e:
            errors.append(f"Could not create master workbook for language '{lang_code}': {e}")
            
    if errors: log_errors_to_file(root_path, errors)
    return len(xliff_files), processed_langs, len(errors)

def import_and_reconstruct_with_glossary(root_path, glossary_path=None):
    input_dir_name = CONFIG["folder_names"]["excel_export"]
    input_dir = root_path / input_dir_name
    if not input_dir.exists(): raise ValueError(f"'{input_dir_name}' folder not found.")
    master_files = list(input_dir.glob('*-master.xlsx'))
    if not master_files: raise ValueError("No '*-master.xlsx' files found.")
    
    errors = []
    try:
        new_glossary_entries = []
        for master_file in master_files:
            lang_code = master_file.name.replace("-master.xlsx", "")
            try:
                df = pd.read_excel(master_file, sheet_name=f"{lang_code}-Translate_Here")
                if 'add_to_glossary' in df.columns:
                    df_to_add = df[df['add_to_glossary'].astype(str).str.strip().str.lower().isin(['x', 'yes'])]
                    for _, row in df_to_add.iterrows():
                        if pd.notna(row['target']) and str(row['target']).strip() != '':
                            new_glossary_entries.append({'source_text': row['source'], 'target_text': str(row['target']), 'language_code': lang_code})
            except ValueError:
                errors.append(f"Skipped glossary update for {master_file.name}: 'Translate_Here' sheet not found.")
        if new_glossary_entries and glossary_path:
            update_glossary_file(glossary_path, new_glossary_entries)
    except Exception as e:
        errors.append(f"A critical error occurred while updating the glossary: {e}")

    translation_map = {}
    for master_file in master_files:
        try:
            lang_code = master_file.name.replace("-master.xlsx", "")
            if lang_code not in translation_map: translation_map[lang_code] = {}
            sheet_name = f"{lang_code}-Translate_Here"
            df = pd.read_excel(master_file, sheet_name=sheet_name)
            if 'target' not in df.columns or 'id_blob' not in df.columns:
                errors.append(f"Missing required columns (target or id_blob) in {master_file.name}.")
                continue
            df.dropna(subset=['target', 'id_blob'], inplace=True)
            for _, row in df.iterrows():
                target_text = str(row['target'])
                blob = str(row['id_blob'])
                ids = decompress_ids(blob)
                for unit_id in ids:
                    translation_map[lang_code][unit_id] = target_text
        except Exception as e:
            errors.append(f"Could not build translation map from {master_file.name}: {e}")

    xliff_output_name = CONFIG["folder_names"]["xliff_output"]
    translated_dir = root_path / xliff_output_name
    separate_lang_dir = translated_dir / "Separate Languages"
    translated_dir.mkdir(exist_ok=True); separate_lang_dir.mkdir(exist_ok=True)
    processed_count = 0
    original_files = list(root_path.glob('*.xliff'))
    
    for xliff_file in original_files:
        try:
            lang_code_xml = get_target_language(xliff_file)
            lang_specific_map = translation_map.get(lang_code_xml, {})
            if not lang_specific_map:
                for key in translation_map:
                    if key.lower() == lang_code_xml.lower():
                        lang_specific_map = translation_map[key]
                        break
            modified_tree = update_xliff_from_map(xliff_file, lang_specific_map)
            primary_path = translated_dir / xliff_file.name
            modified_tree.write(primary_path, pretty_print=True, xml_declaration=True, encoding='UTF-8')
            lang_dir = separate_lang_dir / lang_code_xml
            lang_dir.mkdir(exist_ok=True)
            secondary_path = lang_dir / xliff_file.name
            modified_tree.write(secondary_path, pretty_print=True, xml_declaration=True, encoding='UTF-8')
            processed_count += 1
        except Exception as e:
            errors.append(f"Could not reconstruct file {xliff_file.name}: {e}")
    try:
        process_standard_files(root_path, translated_dir)
    except Exception as e:
        errors.append(f"An error occurred during standard file processing: {e}")

    if errors: log_errors_to_file(root_path, errors)
    return processed_count, len(errors)

# --- XLIFF EDITOR TAB CLASS ---

class XliffEditorTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, padding=10)
        self.tree = None
        self.current_file_path = None
        self.xml_tree = None
        self.namespaces = {'xliff': 'urn:oasis:names:tc:xliff:document:1.2'}
        self.data_store = [] 
        self.current_editing_id = None 

        # --- CONTROLS FRAME (Top) ---
        controls_frame = ttk.Frame(self)
        controls_frame.pack(fill=X, pady=(0, 10))
        
        # File Operations
        ttk.Button(controls_frame, text="📂 Open XLIFF", command=self.load_file, bootstyle="info").pack(side=LEFT, padx=(0, 5))
        ttk.Button(controls_frame, text="💾 Save Changes", command=self.save_file, bootstyle="success").pack(side=LEFT, padx=(0, 20))
        
        # Search & Filter
        ttk.Label(controls_frame, text="Search:").pack(side=LEFT, padx=(0, 5))
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(controls_frame, textvariable=self.search_var, width=25)
        self.search_entry.pack(side=LEFT, padx=(0, 10))
        self.search_entry.bind("<Return>", self.apply_filter)
        self.search_entry.bind("<KeyRelease>", self.apply_filter)
        
        ttk.Label(controls_frame, text="Status:").pack(side=LEFT, padx=(10, 5))
        self.filter_var = tk.StringVar(value="All")
        self.filter_combo = ttk.Combobox(controls_frame, textvariable=self.filter_var, state="readonly", width=15)
        self.filter_combo['values'] = ("All", "New", "Needs Review", "Translated", "Final")
        self.filter_combo.pack(side=LEFT)
        self.filter_combo.bind("<<ComboboxSelected>>", self.apply_filter)

        # --- PANED WINDOW (Split View) ---
        # FIX: Use standard tk_ttk.PanedWindow because ttkbootstrap sometimes misses it
        self.paned_window = tk_ttk.PanedWindow(self, orient=VERTICAL)
        self.paned_window.pack(fill=BOTH, expand=True)

        # 1. TOP PANE: TREEVIEW
        self.tree_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(self.tree_frame, weight=2)

        cols = ("id", "status", "source", "target")
        self.tree = ttk.Treeview(self.tree_frame, columns=cols, show="headings", selectmode="browse")
        
        self.tree.heading("id", text="ID")
        self.tree.heading("status", text="Status")
        self.tree.heading("source", text="Original Source")
        self.tree.heading("target", text="Translated Target")
        
        self.tree.column("id", width=60, stretch=False)
        self.tree.column("status", width=100, stretch=False)
        self.tree.column("source", width=350)
        self.tree.column("target", width=350)
        
        scrollbar = ttk.Scrollbar(self.tree_frame, orient=VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.tree.pack(side=LEFT, fill=BOTH, expand=True)
        
        self.tree.bind("<<TreeviewSelect>>", self.on_row_selected)

        # 2. BOTTOM PANE: EDITOR PANEL
        self.editor_frame = ttk.Labelframe(self.paned_window, text="Edit Selected Segment", padding=10, bootstyle="secondary")
        self.paned_window.add(self.editor_frame, weight=1)

        self.editor_frame.columnconfigure(1, weight=1)
        self.editor_frame.rowconfigure(1, weight=1)
        self.editor_frame.rowconfigure(3, weight=1)

        ttk.Label(self.editor_frame, text="Original Source:", font=("Helvetica", 9, "bold")).grid(row=0, column=0, sticky=W, pady=(0,5))
        # High Contrast Text (White BG, Black Text)
        self.txt_source = tk.Text(self.editor_frame, height=3, bg="white", fg="black", insertbackground="black", state=DISABLED) 
        self.txt_source.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(0, 10))

        ttk.Label(self.editor_frame, text="Translation Target:", font=("Helvetica", 9, "bold")).grid(row=2, column=0, sticky=W, pady=(0,5))
        self.txt_target = tk.Text(self.editor_frame, height=3, bg="white", fg="black", insertbackground="black")
        self.txt_target.grid(row=3, column=0, columnspan=2, sticky="nsew", pady=(0, 10))

        controls_bot = ttk.Frame(self.editor_frame)
        controls_bot.grid(row=4, column=0, columnspan=2, sticky="ew")

        ttk.Label(controls_bot, text="Set Status:").pack(side=LEFT)
        self.edit_status_var = tk.StringVar()
        self.status_dropdown = ttk.Combobox(controls_bot, textvariable=self.edit_status_var, values=("new", "needs-review", "translated", "final"), state="readonly", width=15)
        self.status_dropdown.pack(side=LEFT, padx=10)

        ttk.Button(controls_bot, text="Confirm Update", command=self.update_record, bootstyle="success").pack(side=RIGHT)
        ttk.Button(controls_bot, text="Clear/Revert", command=self.revert_editor, bootstyle="secondary-outline").pack(side=RIGHT, padx=10)

    def load_file(self):
        filepath = filedialog.askopenfilename(title="Select XLIFF File", filetypes=[("XLIFF Files", "*.xliff"), ("All Files", "*.*")])
        if not filepath: return
        
        self.current_file_path = filepath
        try:
            self.xml_tree = etree.parse(filepath)
            self.data_store = []
            
            for trans_unit in self.xml_tree.xpath('//xliff:trans-unit', namespaces=self.namespaces):
                unit_id = trans_unit.get('id')
                
                # --- FIX: Handle Empty Source Tags safely ---
                source_node = trans_unit.find('xliff:source', namespaces=self.namespaces)
                # Use (source_node.text or "") to ensure None becomes ""
                source_text = (source_node.text or "") if source_node is not None else ""
                
                # --- FIX: Handle Empty Target Tags safely ---
                target_node = trans_unit.find('xliff:target', namespaces=self.namespaces)
                if target_node is not None:
                    target_text = target_node.text or ""
                    status = target_node.get('state', 'New') 
                else:
                    target_text = ""
                    status = "New"

                self.data_store.append({
                    "id": unit_id,
                    "source": source_text,
                    "target": target_text,
                    "status": status,
                    "xml_node": trans_unit
                })
            
            self.apply_filter()
            messagebox.showinfo("Loaded", f"Loaded {len(self.data_store)} text blocks.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Could not parse file: {e}")

    def apply_filter(self, event=None):
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        status_filter = self.filter_var.get().lower()
        search_term = self.search_var.get().lower()
        
        for item in self.data_store:
            # Safely get status, defaulting to "new" if missing
            current_status = str(item.get('status', 'new')).lower()
            
            # Check Status Match
            status_match = (status_filter == "all") or (current_status.replace(" ", "") == status_filter.replace(" ", ""))
            
            # Check Search Match (Ensure we search safely on strings)
            s_source = str(item['source'] or "").lower()
            s_target = str(item['target'] or "").lower()
            s_id = str(item['id'] or "").lower()
            
            text_match = True
            if search_term:
                text_match = (search_term in s_source) or (search_term in s_target) or (search_term in s_id)

            if status_match and text_match:
                # --- MULTI-LINE FIX ---
                # Ensure we are operating on strings, not None
                display_source = str(item['source'] or "").replace('\n', ' ').replace('\r', '')
                display_target = str(item['target'] or "").replace('\n', ' ').replace('\r', '')
                
                tag = current_status.replace(" ", "_")
                self.tree.insert("", "end", values=(item['id'], item['status'], display_source, display_target), tags=(tag,))

        # Colors
        self.tree.tag_configure('new', foreground='#ff4d4d') 
        self.tree.tag_configure('needs_review', foreground='#ffad33')
        self.tree.tag_configure('translated', foreground='#33cc33')
        self.tree.tag_configure('final', foreground='#3399ff')

    def on_row_selected(self, event):
        selected_items = self.tree.selection()
        if not selected_items: return
        
        item_values = self.tree.item(selected_items[0], 'values')
        xliff_id = item_values[0]
        
        record = next((r for r in self.data_store if r['id'] == xliff_id), None)
        if record:
            self.load_editor_panel(record)

    def load_editor_panel(self, record):
        self.current_editing_id = record['id']
        
        self.txt_source.config(state=NORMAL)
        self.txt_source.delete("1.0", END)
        self.txt_source.insert("1.0", str(record['source'] or ""))
        self.txt_source.config(state=DISABLED)
        
        self.txt_target.delete("1.0", END)
        self.txt_target.insert("1.0", str(record['target'] or ""))
        
        self.edit_status_var.set(record['status'])

    def revert_editor(self):
        if self.current_editing_id:
             record = next((r for r in self.data_store if r['id'] == self.current_editing_id), None)
             if record: self.load_editor_panel(record)

    def update_record(self):
        if not self.current_editing_id: return

        new_target = self.txt_target.get("1.0", "end-1c")
        new_status = self.edit_status_var.get()
        
        record = next((r for r in self.data_store if r['id'] == self.current_editing_id), None)
        if record:
            record['target'] = new_target
            record['status'] = new_status
            
            trans_unit = record['xml_node']
            target_node = trans_unit.find('xliff:target', namespaces=self.namespaces)
            if target_node is None:
                target_node = etree.SubElement(trans_unit, f"{{{self.namespaces['xliff']}}}target")
            target_node.text = new_target
            target_node.set('state', new_status)
            
            # Refresh View (preserving selection if possible)
            self.apply_filter()
            
            for child in self.tree.get_children():
                if str(self.tree.item(child, 'values')[0]) == str(self.current_editing_id):
                    self.tree.selection_set(child)
                    self.tree.see(child)
                    break

    def save_file(self):
        if not self.current_file_path or not self.xml_tree:
            messagebox.showwarning("Warning", "No file loaded.")
            return
            
        try:
            self.xml_tree.write(self.current_file_path, pretty_print=True, xml_declaration=True, encoding='UTF-8')
            messagebox.showinfo("Success", f"File saved successfully:\n{self.current_file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save file: {e}")

# --- MAIN GUI CLASS ---

class FinalConverterApp(ttk.Window):
    def __init__(self):
        super().__init__(themename="superhero")
        self.title("Translation Project Tool")
        self.geometry("800x700")
        self.glossary_path = None
        
        # --- HEADER ---
        header_frame = ttk.Frame(self, padding=10)
        header_frame.pack(fill=X)
        ttk.Label(header_frame, text="Localization Toolkit", font=("Helvetica", 16, "bold")).pack(side=LEFT)
        ttk.Button(header_frame, text="⚙ Config", command=self.open_config, bootstyle="outline-secondary").pack(side=RIGHT)

        # --- NOTEBOOK (TABS) ---
        self.notebook = ttk.Notebook(self, bootstyle="primary")
        self.notebook.pack(fill=BOTH, expand=True, padx=10, pady=10)

        # TAB 1: CONVERTER
        self.tab_converter = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(self.tab_converter, text="Converter Tools")
        self.build_converter_tab()

        # TAB 2: XLIFF EDITOR
        self.tab_editor = XliffEditorTab(self.notebook)
        self.notebook.add(self.tab_editor, text="XLIFF Editor")

        # --- STATUS BAR ---
        self.status_frame = ttk.Frame(self, padding=10)
        self.status_frame.pack(side=BOTTOM, fill=X)
        self.progress = ttk.Progressbar(self.status_frame, mode='indeterminate', bootstyle="success-striped")
        self.status_label = ttk.Label(self.status_frame, text="Ready", font=("Helvetica", 9))
        self.status_label.pack(side=LEFT)

        self.auto_load_glossary()

    def build_converter_tab(self):
        lc = ttk.Labelframe(self.tab_converter, text="1. Analysis & Setup", padding=10, bootstyle="info")
        lc.pack(fill=X, pady=5)
        col1 = ttk.Frame(lc)
        col1.pack(fill=X)
        self.glossary_label = ttk.Label(col1, text="No glossary loaded", foreground="orange")
        self.glossary_label.pack(side=LEFT, padx=5)
        ttk.Button(col1, text="Load Glossary", command=self.load_glossary, bootstyle="secondary-sm").pack(side=RIGHT)
        ttk.Separator(lc, orient=HORIZONTAL).pack(fill=X, pady=10)
        ttk.Button(lc, text="Analyze Project Statistics", command=lambda: self.start_thread(self.run_analysis), bootstyle="info").pack(fill=X)

        l1 = ttk.Labelframe(self.tab_converter, text="2. Export for Translation", padding=10, bootstyle="primary")
        l1.pack(fill=X, pady=10)
        ttk.Button(l1, text="Create Excel Masters (Step 1)", command=lambda: self.start_thread(self.run_export), bootstyle="primary").pack(fill=X, pady=2)
        ttk.Button(l1, text="Apply DeepL Translations (Step 1.5)", command=lambda: self.start_thread(self.run_apply_deepl), bootstyle="primary-outline").pack(fill=X, pady=2)

        l2 = ttk.Labelframe(self.tab_converter, text="3. Import & Reconstruct", padding=10, bootstyle="success")
        l2.pack(fill=X, pady=5)
        ttk.Button(l2, text="Reconstruct XLIFFs (Step 2)", command=lambda: self.start_thread(self.run_import), bootstyle="success").pack(fill=X)

    def start_thread(self, target_func):
        self.progress.pack(side=RIGHT, fill=X, expand=True, padx=10)
        self.progress.start(10)
        self.status_label.config(text="Processing... Please wait.")
        thread = threading.Thread(target=self.run_wrapper, args=(target_func,))
        thread.start()

    def run_wrapper(self, func):
        try: func() 
        finally: self.after(0, self.stop_progress)

    def stop_progress(self):
        self.progress.stop()
        self.progress.pack_forget()
        self.status_label.config(text="Ready")

    def open_config(self):
        config_path = Path("config.json")
        if config_path.exists(): os.startfile(config_path)
        else: messagebox.showwarning("Missing", "config.json not found.")

    def run_apply_deepl(self):
        root_dir = filedialog.askdirectory(title="Select Root Folder")
        if not root_dir: return
        def worker():
            try:
                updated, total, errors = apply_deepl_translations(Path(root_dir))
                msg = f"Updated {updated}/{total} files."
                if errors: 
                    log_errors_to_file(Path(root_dir), errors)
                    msg += " Check logs."
                messagebox.showinfo("Result", msg)
            except Exception as e:
                messagebox.showerror("Error", str(e))
        self.start_thread(worker)

    def run_export(self):
        root_dir = filedialog.askdirectory(title="Select Root Folder")
        if not root_dir: return
        def worker():
            try:
                fc, lc, ec = export_to_excel_with_glossary(Path(root_dir), self.glossary_path)
                messagebox.showinfo("Result", f"Processed {fc} files ({lc} langs). Errors: {ec}")
            except Exception as e:
                messagebox.showerror("Error", str(e))
        self.start_thread(worker)

    def run_import(self):
        root_dir = filedialog.askdirectory(title="Select Root Folder")
        if not root_dir: return
        def worker():
            try:
                pc, ec = import_and_reconstruct_with_glossary(Path(root_dir), self.glossary_path)
                messagebox.showinfo("Result", f"Reconstructed {pc} files. Errors: {ec}")
            except Exception as e:
                messagebox.showerror("Error", str(e))
        self.start_thread(worker)

    def run_analysis(self):
        root_dir = filedialog.askdirectory(title="Select Root Folder")
        if not root_dir: return
        def worker():
            try:
                data = perform_analysis(Path(root_dir), self.glossary_path)
                self.after(0, lambda: self.display_analysis_report(data))
            except Exception as e:
                messagebox.showerror("Error", str(e))
        self.start_thread(worker)

    def display_analysis_report(self, data):
        report_window = ttk.Toplevel(self)
        report_window.title("Translation Analysis Report")
        report_window.geometry("800x500")
        
        top_frame = ttk.Frame(report_window, padding=10)
        top_frame.pack(fill=X)
        ttk.Button(top_frame, text="Export to TXT", command=lambda: self.export_report_to_text(data)).pack()
        
        cols = ['Language', 'Total Words', 'Repetitions', 'Glossary Matches', 'New Words']
        tree = ttk.Treeview(report_window, columns=cols, show="headings", bootstyle="info")
        
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=120, anchor='center')
        
        totals = {key: 0 for key in ['Total Words', 'Repetitions', 'Glossary Matches', 'New Words']}
        for lang, metrics in data.items():
            row = (lang, metrics['Total Words'], metrics['Repetitions'], metrics['Glossary Matches'], metrics['New Words'])
            tree.insert("", "end", values=row)
            for key in totals:
                totals[key] += metrics.get(key, 0)
        
        tree.insert("", "end", values=())
        total_row = ('TOTAL', totals['Total Words'], totals['Repetitions'], totals['Glossary Matches'], totals['New Words'])
        tree.insert("", "end", values=total_row, tags=('totalrow',))
        tree.tag_configure('totalrow', font=('Helvetica', 10, 'bold'))
        tree.pack(expand=True, fill="both", padx=10, pady=10)

    def export_report_to_text(self, data):
        filepath = filedialog.asksaveasfilename(title="Save Report As", defaultextension=".txt", filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if not filepath: return
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                headers = ['Language', 'Total Words', 'Repetitions', 'Glossary Matches', 'New Words']
                widths = [20, 15, 15, 18, 15]
                header_line = "".join(h.ljust(w) for h, w in zip(headers, widths))
                f.write(f"{header_line}\n")
                f.write(f"{'-' * sum(widths)}\n")
                totals = {key: 0 for key in ['Total Words', 'Repetitions', 'Glossary Matches', 'New Words']}
                for lang, metrics in data.items():
                    row_values = [lang, metrics.get('Total Words', 0), metrics.get('Repetitions', 0), metrics.get('Glossary Matches', 0), metrics.get('New Words', 0)]
                    f.write("".join(str(v).ljust(w) for v, w in zip(row_values, widths)) + "\n")
                    for i, key in enumerate(totals):
                        totals[key] += row_values[i+1]
                f.write(f"{'-' * sum(widths)}\n")
                total_values = ['TOTAL', totals['Total Words'], totals['Repetitions'], totals['Glossary Matches'], totals['New Words']]
                f.write("".join(str(v).ljust(w) for v, w in zip(total_values, widths)) + "\n")
            messagebox.showinfo("Success", f"Report successfully saved to:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not save the report: {e}")

    def auto_load_glossary(self):
        default_path = Path("glossary.xlsx")
        if default_path.exists(): self.set_glossary(default_path)

    def load_glossary(self):
        filepath = filedialog.askopenfilename(title="Select Glossary Excel File", filetypes=[("Excel files", "*.xlsx")])
        if filepath: self.set_glossary(filepath)

    def set_glossary(self, path):
        self.glossary_path = path
        filename = Path(path).name
        self.glossary_label.config(text=f"Using: {filename}", foreground="green")

if __name__ == "__main__":
    app = FinalConverterApp()
    app.mainloop()
